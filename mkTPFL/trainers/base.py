from fedAvg.trainers.base import BaseTrainer
from fedAvg.models.model import choose_model
from fedAvg.optimizers.gd import GD
from fedAvg.utils.worker_utils import mkdir

from math import ceil
from sympy import I
from collections import defaultdict
from tqdm import tqdm
import os
import sys
import time
import torch
from torch.utils.data import DataLoader
import pickle
import random
from math import ceil, floor
from HEAAN import Ciphertext, Ciphertexts
import numpy
import openpyxl


from mkTPFL.ckks.ckks_parameters import CKKSParameters, print_parameters
from mkTPFL.ckks.ckks_decryptor import CKKSDecryptor
from mkTPFL.ckks.ckks_encoder import CKKSEncoder, CKKSMessage
from mkTPFL.ckks.ckks_encryptor import CKKSEncryptor
from mkTPFL.ckks.ckks_evaluator import CKKSEvaluator
from mkTPFL.ckks.ckks_key_generator import CKKSKeyGenerator

from mkTPFL.mklhs.mklhs_parameters import MKLHSParameters, print_mklhs_parameters
from mkTPFL.utils.lhh_parameters import LHH_init

from mkTPFL.roles.dataisland import DataIsland
from mkTPFL.roles.flserver import FLServer
# from mkTPFL.roles.model_evaluator import EncModelEvaluator, ValiDataGenerator
from mkTPFL.roles.model_evaluator import EncModelEvaluator, preprocessing, unmask

from mkTPFL.utils.class_transformer import *
from mkTPFL.utils.info_util import role2Data, data2Role, data2EvalDatasets, saveDIKeys, setDIKeyFiles, saveSVKeys, setSVKeyFiles, build_trainer_from_file

from pympler import asizeof

class BaseClient():
    def __init__(self, cid, weight_info=(None,400,700)):
        self.cid = cid
        self.weight = 1
        seed, w_min, w_max = weight_info
        # print('weight_info:', weight_info)
        if seed or seed==0:
            np.random.seed(seed+cid)
            self.weight = np.random.randint(w_min, w_max)


class BaseMKTPFLTrainer(BaseTrainer):
    def __init__(self, basic_options, dataset, verify_data=None, ckks_params=None, mklhs_params=None):
        options = defaultdict(lambda:False)
        options.update(basic_options)

        self.dataset_name = options['dataset']
        self.model_name = options['model']
        self.seed = options['seed']

        if self.dataset_name == 'randomvector':
            self.model = options['input_length']
            self.model_len = options['input_length']
            self.input_length = options['input_length']
            self.num_round = options['num_round']
            self.clients_per_round = options['clients_per_round']
            self.simple_average = not options['noaverage']
            w_seed = self.seed if not self.simple_average else None
            weight_info = (w_seed, options['weight_min'], options['weight_max'])
            self.clients = [ BaseClient(cid, weight_info=weight_info) for cid in range(options['clients_num']) ]
        else:
            self.model = choose_model(options)
            self.move_model_to_gpu(self.model, options)
            self.optimizer = GD(self.model.parameters(), lr=options['lr'], weight_decay=options['wd'])
            super(BaseMKTPFLTrainer, self).__init__(options, dataset, self.model, self.optimizer)
            self.model_len = len(self.worker.get_flat_model_params().detach())
            self.clients = self.clients[:options['clients_num']]

        # for save setup or local data
        self.build_from_file = options['build_from_file']
        self.setup_data_dir = mkdir(options['setup_data_dir'])
        self.trainer_pfl_info_file = options['trainer_pfl_info_file']
        self.local_data_dir = options['local_data_dir']
        (self.traindata_path, self.traindata_file_key) = options['traindata_info']
        self.val_c_flag = options['eval_client']
        if self.val_c_flag:
            self.eval_method = options['eval_method']
            if verify_data:
                self.verify_dataloader = DataLoader(verify_data, batch_size=self.batch_size, shuffle=False)
            else:
                self.verify_dataloader = None
        self.stat_dir = options['stat_dir']
        
        self.key_files_flag = options['key_files']
        self.dikeys_dir = options['dikeys_dir']
        self.svkeys_dir = options['svkeys_dir']
        self.diskshares_dir = options['diskshares_dir']
        self.save_pfl_flag = options['save_pfl']
        self.eval_mem_flag = options['eval_mem']

        # for PFL: params, dis={cid:di}, svs=[sv]
        # set number of DIs and SVs
        self.di_num = len(self.clients)
        self.sv_num = 1 if not options['sv_num'] else options['sv_num']
        if self.sv_num == 1:
            self.di_threshold = self.di_num if not options['clients_threshold'] else min(options['clients_threshold'], self.di_num)
            self.sv_threshold = self.sv_num
        else:
            self.di_threshold = self.di_num
            self.sv_threshold = self.sv_num if not options['sv_threshold'] else min(options['sv_threshold'], self.sv_num) 

        # set public parameters
        self.set_clients_weight() # set weights of clients
        self.soln_scaling = options['soln_scaling'] if options['soln_scaling'] else 10**6
        self.dropout_ratio = options['dropout_ratio']
        self.dropout_flag = True if self.dropout_ratio>0 else False
        self.mklhs_params = mklhs_params if mklhs_params else MKLHSParameters()
        # ckks_params
        self.ckks_params = ckks_params

        # setup trainer from file or not
        if self.build_from_file: 
            # setup ckks_params from file
            if not self.ckks_params: 
                file_path = self.setup_data_dir+'ckks_params_data.pkl'
                if os.path.exists(file_path):
                    data = self.trainer_get_data(file_path=file_path, print_flag=True)
                    self.ckks_params = data2Role(data, 'CKKSParams')
                    print('>>> set ckks_parameters from file [{}]'.format(file_path))
                else:
                    print('>>> no ckks_parameters file [{}], setting new CKKSParams ...'.format(file_path))
                    self.build_from_file = False
                    self.ckks_params = CKKSParameters()
                    print_parameters(self.ckks_params)
                    file_path = self.setup_data_dir+'ckks_params_data.pkl'
                    data = role2Data(self.ckks_params, 'CKKSParams')
                    self.trainer_save_data(data, file_path=file_path, print_flag=True)
                    self.build_new_trainer()
            build_trainer_from_file(self)

        else:
            self.build_new_trainer()

        # print_parameters(self.ckks_params)
        print('>>> Model [{}] params: {}'.format(self.model_name, self.model_len))
        print('>>> Clients\' weights:\n\t', self.clients_weights)
        print('>>> Build MKTPFLTrainer Done!!')
        

    def build_new_trainer(self):
        ''' setup new trainer '''
        print('>>> building new Trainer ...')
        self.setup_singlesv() if self.sv_num == 1 else self.setup_multisv()
        if self.save_pfl_flag:
            print('Saving PFL information ...')
            # trainer_info
            dis_info, svs_info = {}, []
            for (cid, di) in self.dis.items():
                dis_info[cid] = role2Data(di,'DI')
            for sv in self.svs:
                svs_info.append( role2Data(sv,'SV') )
            trainer_pfl_info = { 'dis': dis_info, 'svs': svs_info }
            file_path = self.trainer_pfl_info_file+'_trainer_info_data.pkl'
            self.trainer_save_data(trainer_pfl_info, file_path=file_path, print_flag=True)


    def set_clients_weight(self, acc=4):
        clients_weights = None
        if not self.simple_average:
            clients_weights = defaultdict(int)
            max_w = 10**(acc)
            for c in self.clients:
                try:
                    w = len(c.train_data)
                except:
                    w = c.weight
                clients_weights[c.cid] = w if w<max_w else round( int(str(w)[:acc+1])/10 )
        self.clients_weights = clients_weights
        # self.clients_weights = [2]*len(self.clients)


    def setup_singlesv(self):
        sv_num, di_num = self.sv_num, self.di_num
        dikeys_dir = mkdir(self.dikeys_dir)
        di_public_key_list, di_enc_public_key_list, di_enc_relin_key_list, di_enc_rot_keys_list, di_sign_public_key_list = [], [], [], [], []

        # Initialize DataIslands (DIs)
        print('Initialize DataIslands: ')
        di_inx = 0
        dis = {}
        update_key_flag = False
        for c in self.clients:
            di = DataIsland(di_inx, di_num=di_num, sv_num=sv_num, ckks_params=self.ckks_params, di_threshold=self.di_threshold)
            print(' DataIsland ['+str(di_inx)+'] initializing ...')
            di_inx += 1
            # Generate the setup secret key and pubilc key of DataIslands
            existed_key_file = False
            if self.build_from_file:
                existed_key_file = setDIKeyFiles(di, dikeys_dir, print_flag=True)
            if not existed_key_file:
                di.keyGen(eval_flag=self.val_c_flag)
                update_key_flag = True
            if not existed_key_file and self.key_files_flag:
                saveDIKeys(di, dikeys_dir, print_flag=True, reset_flag=True)
            dis[c.cid] = di
            if self.key_files_flag:
                di_public_key_list.append(di.public_key_file)
                di_enc_public_key_list.append(di.enc_public_key_file)
                di_enc_relin_key_list.append(di.enc_relin_key_file)
                di_enc_rot_keys_list.append(di.enc_rot_keys_file)
                di_sign_public_key_list.append(di.sign_public_key_file)
            else:
                di_public_key_list.append(di.public_key)
                di_enc_public_key_list.append(di.enc_public_key)
                di_enc_relin_key_list.append(di.enc_relin_key)
                di_enc_rot_keys_list.append(di.enc_rot_keys)
                di_sign_public_key_list.append(di.sign_public_key)
        # Set the di's di_public_keys
        for di in dis.values():
            di.set_di_public_keys(di_public_key_list)
            di.set_di_sign_public_keys(di_sign_public_key_list)
        self.dis = dis
        if self.di_threshold < self.di_num:
            # print('update_key_flag:', update_key_flag)
            self.gen_di_sk_shares(update_flag=update_key_flag)
        
        # Initialize Server (SV)
        print('Initialize Server')
        sv = FLServer(0, di_num=di_num, sv_num=sv_num, ckks_params=self.ckks_params, di_threshold=self.di_threshold)
        sv.set_di_public_keys(di_public_key_list)
        sv.set_di_enc_public_keys(di_enc_public_key_list)
        sv.set_di_enc_relin_keys(di_enc_relin_key_list)
        sv.set_di_enc_rot_keys(di_enc_rot_keys_list)
        sv.set_di_sign_public_keys(di_sign_public_key_list)
        # save the svs' keys into file
        if self.build_from_file:
            existed_key_file = setSVKeyFiles(sv, self.svkeys_dir, print_flag=True)
        if not existed_key_file and self.key_files_flag:
            saveSVKeys(sv, self.svkeys_dir)
        self.svs = [sv]

    def gen_di_sk_shares(self, update_flag=False): 
        ''' Generate DIs' secret key shares (save in file) '''
        dir_path = mkdir(self.diskshares_dir)
        for di in self.dis.values():
            di_inx =  di.di_index
            di_dir_path = dir_path+'DI_'+str(di_inx)+'/'
            if os.path.exists(di_dir_path):
                di.di_secret_key_shares_dir = di_dir_path
            else:
                print('>>> no sk_shares dir [{}].'.format(di_dir_path))
                update_flag = True
                break

        if update_flag and self.di_threshold<self.di_num:
            # init di.di_secret_key_shares_file
            for di in self.dis.values():
                di.di_secret_key_shares_dir = mkdir(self.diskshares_dir+'DI_'+str(di.di_index)+'/')
            print('DIs\' secret keys sharing to DIs  ...')
            for di in self.dis.values():
                di_inx =  di.di_index
                secret_key_shares = di.secretkeySharesGen()
                # save to file
                begin_time = time.time()
                for dj in self.dis.values():
                    dj_inx =  dj.di_index
                    sk_share_data = DataKeyShare(secret_key_shares[dj_inx])
                    dj_file_path = dj.di_secret_key_shares_dir + str(di_inx)+'.pkl'
                    self.trainer_save_data(sk_share_data, file_path=dj_file_path, print_flag=False)
                print('  >>> save DI['+str(di_inx)+'] secret key shares, using '
                    +str(round(time.time()-begin_time, 4))+' seconds')
                del secret_key_shares

        # generate agg_sk_share (default)
        for di in self.dis.values():
            di_inx =  di.di_index
            di_file_path = dir_path+'DI_'+str(di_inx)+'/'+'agg.pkl'
            begin_time = time.time()
            if not os.path.exists(di_file_path):
                agg_sk_share_data = DataKeyShare(di.get_agg_sk_share())
                self.trainer_save_data(agg_sk_share_data, file_path=di_file_path, print_flag=False)
                print('  >>> save DI['+str(di_inx)+'] aggregation secret key share, using '
                    +str(round(time.time()-begin_time, 4))+' seconds')
                del agg_sk_share_data


    def set_agg_pk(self, selected_clients=None):
        ''' generate the aggregation_public_key '''
        print('(DIs & SV) set the aggregation public key...', end='', flush=True)
        selected_clients = selected_clients if selected_clients else self.clients
        di_index_list = [ self.dis[c.cid].di_index for c in selected_clients ]
        begin_time = time.time()
        agg_pk = self.svs[0].aggpublickeyGen(di_index_list)
        for di in self.dis.values():
            di.aggregation_public_key = agg_pk
        for sv in self.svs:
            sv.aggregation_public_key = agg_pk
        time_agg_pk = round(time.time()-begin_time, 4)
        print('\t done! using {} seconds'.format(time_agg_pk))
        return agg_pk


    def local_encrypt_iterkey(self, selected_clients, eval_mem_flag=False, save_flag=False, round_num='tmp'):
        ''' each DI: 
            1. partciph_s = partEncrypt(enc_serect_key)
            2. sign_cps = sign(partciph_s)

            output: partciph_s_dict, sign_cps_dict, stats
                - partciph_s_dict = {cid: partciph_s}
                    partciph_s = partciph_s_file if save_flag else partciph_s
                - sign_cps_dict = {cid:sign_cps}
                    partciph_s = partciph_s_file if save_flag else partciph_s
                - stats[cid] includes 'time_enc_sk', 'bytes_enc_sk', 'time_sign_cps', 'bytes_enc_cps'
        '''
        # stats
        stats = defaultdict(dict)
        eval_mem_flag = eval_mem_flag if eval_mem_flag else self.eval_mem_flag
        
        # Get the aggregation public key
        print('(DIs) PartEncrypt iterkey & Sign ciph_iterkey:')
        partciph_s_dict = {} # part ciphertexts of enc_serect_key
        sign_cps_dict = {} # signatures of partciph_ss, sign_cps_dict[cid] = sign(partciph_s)

        log_flag = False # only log one client
        for c in tqdm(selected_clients):
            cid = c.cid
            di = self.dis[cid]
            # partEncrypt(enc_serect_key)
            tqdm.write('\n\033[F >> DI['+str(di.di_index)+'] (cid:'+str(cid)+')'+'\t encrypting iterkey...')
            begin_time = time.time()
            iterkey = di.iterkeyGen()
            partciph_s = di.partEncIterkey(iterkey)
            partciph_s_dict[cid] = partciph_s
            end_time = time.time()
            if not log_flag:
                stats[cid].update({'time_enc_sk':round(end_time-begin_time, 4)})
            # tqdm.write('', end='\r')

            # signature ciphertexts of part_iterkey
            tqdm.write('\033[F >> DI['+str(di.di_index)+'] (cid:'+str(cid)+')'+'\t signing iterkey...   ')
            begin_time = time.time()
            sign_cps = di.signCiphtxts(partciph_s)
            sign_cps_dict[cid] = sign_cps
            end_time = time.time()
            if not log_flag:
                stats[cid].update({'time_sign_cps':round(end_time-begin_time, 4)})
            log_flag = True

            if save_flag: # save to file
                partciph_s_data = DataCiphertexts(partciph_s, print_flag=True)
                di_dir = mkdir(self.local_data_dir+'DI_'+str(di.di_index)+'/')
                di_file_path = di_dir +  'round_' + str(round_num) + '_ciph_s_data.pkl'
                self.trainer_save_data(partciph_s_data, file_path=di_file_path, print_flag=True)
                del partciph_s_dict[cid]
                partciph_s_dict[cid] = di_file_path
            tqdm.write('\033[F', end='\r') # remove the prompt in tqdm

        if eval_mem_flag: # statistics memory size
            cid = selected_clients[0].cid # sample client
            di = self.dis[cid] 
            partciph_s_data = DataCiphModel(partciph_s_dict[cid], print_flag=True)
            di_dir = mkdir(self.local_data_dir+'DI/')
            di_file_path = di_dir + 'ciph_s_data.pkl'
            self.trainer_save_data(partciph_s_data, file_path=di_file_path, print_flag=True)
            stats[cid].update({'bytes_enc_sk': os.path.getsize(di_file_path)})
            # print('ByteSize of partciph_s:', partciph_s_dict[cid].ByteSize())
            print('getsizeof of partciph_s:', sys.getsizeof(partciph_s_dict[cid]))
            print('asizeof of partciph_s:', asizeof.asizeof(partciph_s_dict[cid]))
            print('getsizeof of partciph_s_data:', sys.getsizeof(partciph_s_data))
            print('asizeof of partciph_s_data:', asizeof.asizeof(partciph_s_data))
            print('bytes_enc_sk:', os.path.getsize(di_file_path))

            sign_cps_data = sign_cps_dict[cid].getSignStr()
            di_dir = mkdir(self.local_data_dir+'DI/')
            di_file_path = di_dir + 'sign_cps_data.pkl'
            self.trainer_save_data(sign_cps_data, file_path=di_file_path, print_flag=True)
            stats[cid].update({'bytes_sign_cps': os.path.getsize(di_file_path)})

        return partciph_s_dict, sign_cps_dict, stats


    def local_encrypt_soln(self, selected_clients, solns, eval_mem_flag=False, save_flag=False, round_num='tmp'):
        ''' each DI: 
            0. rescale: lm = [round(float(v)*self.soln_scaling) for v in soln]
            1. sign_lm = signModel(soln) = (sign(LHH(lm)), LHH(lm))
            2. ciph_lm = encrypt(lm)
            3. sign_clm = sign(ciph_lm)

            output: sign_lm_dict, ciph_lm_dict, sign_clm_dict, stats
                - sign_lm_dict = {cid: sign_lm} = {cid: sign_h_i}
                    sign_lm = sign_lm_file if save_flag else sign_lm
                }
                - ciph_lm_dict = {cid:ciph_lm}
                    ciph_lm = ciph_lm_file if save_flag else ciph_lm
                - sign_clm_dict = {cid:sign_clm}
                    sign_clm = sign_clm_file if save_flag else sign_clm
                - stats[cid] includes 'time_enc_lm', 'time_sign_lm', 
                            'bytes_lm', 'bytes_enc_lm', 'bytes_sign_lm'
        '''
        # stats
        stats = defaultdict(dict)
        eval_mem_flag = eval_mem_flag if eval_mem_flag else self.eval_mem_flag

        print('(DIs) Sign & Encrypt local_model & Sign ciph_lm:')
        sign_lm_dict = {} # signatures of LHH(lm)s, sign_lm_dict[cid] = (sign(LHH(lm)), LHH(lm))
        ciph_lm_dict = {} # ciphertexts of local models, ciph_lm_dict[cid] = local model ciphertxt
        # sign_clm_dict = {} # signatures of ciph_lms, sign_clm_dict[cid] = sign(ciph_lm)

        log_flag = False # only log one client
        for ci in tqdm(range(len(selected_clients))):
            cid, soln = selected_clients[ci].cid, solns[ci][1]
            di, lm = self.dis[cid], [round(float(v)*self.soln_scaling) for v in soln]
            # print("!!soln_length:", len(lm))
            stats[cid].update({'bytes_lm':sys.getsizeof(lm)})

            # signModel(lm): sign_lm = (sign(LHH(lm)), LHH(lm))
            # print('\tsigning local model ...')
            tqdm.write('\n\033[F >> DI['+str(di.di_index)+'] (cid:'+str(cid)+')'+'\t signing local model...' )
            begin_time = time.time()
            sign_lm = di.signModel(self.model, lm)
            sign_lm_dict[cid] = sign_lm
            end_time = time.time()
            if not log_flag:
                stats[cid].update({'time_sign_lm':round(end_time-begin_time, 4)})
            # # verify test
            # sign_hlm, msg_hlm = sign_lm
            # ver_sign_res = self.svs[0].verSign(di.di_index, msg_hlm, sign_hlm)
            # print('ver_sign_res: ', ver_sign_res)

            # encrypt(lm)
            tqdm.write('\033[F >> DI['+str(di.di_index)+'] (cid:'+str(cid)+')'+'\t encrypting local model...' )
            begin_time = time.time()
            ciph_lm = di.encryptModel(self.model, lm)
            end_time = time.time()
            ciph_lm_dict[cid] = ciph_lm
            if not log_flag:
                stats[cid].update({'time_enc_lm':round(end_time-begin_time, 4)})
            log_flag = True
            # # correctness test
            # lm = di.decryptModel(self.model, ciph_lm)
            # print('dec_w (in encrypt_lm): ', torch.Tensor(lm['weight']))
            # print('dec_b (in encrypt_lm): ', torch.Tensor(lm['bias']))

            if save_flag: # save to file
                ciph_lm_data = DataCiphModel(ciph_lm, print_flag=True)
                di_dir = mkdir(self.local_data_dir+'DI_'+str(di.di_index)+'/')
                di_file_path = di_dir +  'round_' + str(round_num) + 'ciph_lm_data.pkl'
                self.trainer_save_data(ciph_lm_data, file_path=di_file_path, print_flag=True)
                del ciph_lm_dict[cid]
                ciph_lm_dict[cid] = di_file_path

            tqdm.write('\033[F', end='\r') # remove the prompt in tqdm

        if eval_mem_flag: # statistics memory size
            cid = selected_clients[0].cid # sample client
            di = self.dis[cid] 
            ciph_lm_data = DataCiphModel(ciph_lm_dict[cid], print_flag=True)
            di_dir = mkdir(self.local_data_dir+'DI/')
            di_file_path = di_dir + 'ciph_lm_data.pkl'
            self.trainer_save_data(ciph_lm_data, file_path=di_file_path, print_flag=True)
            stats[cid].update({'bytes_enc_lm': os.path.getsize(di_file_path)})

            sign_clm, hash_lm = sign_lm_dict[cid]
            sign_lm_data = (sign_clm.getSignStr(), LHH_init().getHashesHex(hash_lm))
            di_dir = mkdir(self.local_data_dir+'DI/')
            di_file_path = di_dir + 'sign_lm_data.pkl'
            self.trainer_save_data(sign_lm_data, file_path=di_file_path, print_flag=True)
            stats[cid].update({'bytes_sign_lm': os.path.getsize(di_file_path)})

        return sign_lm_dict, ciph_lm_dict, stats


    def evaluation_localmodel(self, cid, ciph_lm, eval_method=None):
        ''' SV & a DI: evaluates the ciph_lm 
                1. (setup) SV holds (rk,pk) of ciph_lm
                2. (offline) SV: compute eval_data and mask_data for SE
                    eval_method: eval_data = 
                        'l2': None;
                        'ln': valid_soln, ciph_soln;
                        'zeno': valid_soln, a, b, ciph_soln;
                        'cos': valid_soln, valid_soln_vec;
                    mask_data = (r1, r2, ciph_r2)
                3. (online) SV: evaluate and mask ciph_lm
                    ciph_eval = eval_method(ciph_lm, eval_data)
                    ciph_mask_eval = r1*ciph_eval+Enc(r2)
                4. (online) DI: decrypt ciph_mask_eval
                    mask_eval = Dec(ciph_mask_eval)
                5. (online) SV: unmask mask_eval
            
            output: eval_res, eval_data, sv_stats, di_stats
        '''
        # stats
        sv_stats = defaultdict(dict)
        di_stats = defaultdict(dict)

        eval_method = eval_method if eval_method else self.eval_method
        print('(SV) Evaluate local model ciphertexts:')
        sv = self.svs[0]
        di = self.dis[cid]
        di_inx = di.di_index

        # (setup) key generation&distribution, vaild_data generation&preprocessing
        sk, pk = di.enc_secret_key, di.enc_public_key
        # relin_key
        try:
            rk = di.enc_relin_key
        except:
            key_generator = CKKSKeyGenerator(self.ckks_params)
            rk = key_generator.generate_relin_key(sk, fixedax_flag=True)
        model_evaluator = EncModelEvaluator(self.ckks_params, pk, rk, eval_method=eval_method, soln_scaling=self.soln_scaling)

        # (offline) SV: compute eval_data for SE
        valid_soln = torch.tensor(self.latest_model)
        if eval_method == 'ln' or eval_method == 'zeno':
            if self.verify_dataloader:
                valid_soln,_ = self.worker.local_train(self.verify_dataloader)

        begin_time = time.time()
        eval_data = model_evaluator.preprocessing(self.worker, self.verify_dataloader, valid_soln=valid_soln, ciph_model=ciph_lm)
        sv_stats['time_eval_data'] = round(time.time()-begin_time, 4)

        # (offline) SV: compute mask_data
        begin_time = time.time()
        mask_data = model_evaluator.getMask()
        sv_stats['time_get_mask'] = round(time.time()-begin_time, 4)

        # (online) SV: evaluate ciph_lm
        begin_time = time.time()
        ciph_eval = model_evaluator.eval(ciph_lm, eval_data)
        sv_stats['time_eva_clm'] = round(time.time()-begin_time, 4)

        # (online) SV: mask ciph_eval
        begin_time = time.time()
        ciph_mask_eval = model_evaluator.mask(ciph_eval, mask_data)
        sv_stats['time_mask_eva'] = round(time.time()-begin_time, 4)

        # (online) DI: decrypt ciph_mask_eval
        begin_time = time.time()
        mask_eval = di.decrypt(ciph_mask_eval, msg_flag=False)
        # print('mask_eval:', mask_eval)
        di_stats['time_dec_eva'] = round(time.time()-begin_time, 4)
        sum_mask_eval = 0
        for i in range(self.ckks_params.slots):
            sum_mask_eval += round(float(str(mask_eval[i].real())), 5)
        

        # (online) SV: unmask mask_eval
        begin_time = time.time()
        r1, r2, ciph_r2 = mask_data
        # print('r2:', r2[:3])
        eval_res = model_evaluator.unmask(sum_mask_eval, mask_data, eval_data=eval_data)
        sv_stats['time_unmask'] = round(time.time()-begin_time, 4)
        
        return eval_res, eval_data, sv_stats, di_stats



    def verification_localsign(self, partciph_s_dict, sign_cps_dict, sign_lm_dict, selected_clients=None, eval_mem_flag=False):
        ''' SV: 
                1. verifies (partciph_s_dict, sign_cps_dict)
                    - ver_sign_cps_dict = {cid: ver_sign_cps}
                    - msg_ciph_s_dict = {cid: msg(ciph_lm)}
                2. verifies sign_lm_dict, and sign_lm_dict[cid] = (sign(LHH(lm)), LHH(lm))
                    - ver_sign_lm_dict = {cid: ver_sign_lm}
                3. get verification result: 
                    ver_sign_dict[cid] = ver_sign_cps_dict[cid] && ver_sign_lm_dict[cid]
            
            output: ver_sign_cps_res, ver_sign_dict, msg_ciph_s_dict, stats
                - ver_sign_cps_res = True if all ver_sign_cps is True else False
                    ( also ver_sign_cps_res = False if any selected_client's partciph_s is missing )
                - ver_sign_dict = {cid: ver_sign_lm && ver_sign_cps}
                - msg_ciph_s_dict = {cid: msg_ciph_s}
                - stats includes 'time_ver_sign_cps', 'time_ver_sign_lm', 
                                'time_ver_local_sign', 'bytes_msg_ciph_s_dict'
        '''
        sv = self.svs[0]
        eval_mem_flag = eval_mem_flag if eval_mem_flag else self.eval_mem_flag
        # stats = {}
        time_ver_sign_lm, time_ver_sign_clm, time_ver_sign_cps = 0, 0, 0
        ver_sign_cps_dict, msg_ciph_s_dict, ver_sign_cps_res = defaultdict(lambda:False), {}, (True,-1)
        ver_sign_lm_dict = defaultdict(lambda:False)
        
        print('(SV) Verify the signatures of Enc(iterkey)s, LHH(lm)s:')
        # Verify signatures of part_iterkeys ciphertext
        begin_time = time.time()
        selected_cids = [ c.cid for c in selected_clients ] if selected_clients else list(partciph_s_dict.keys())
        # verify signatures of part_iterkeys ciphertext
        
        for cid in selected_cids:
            if cid not in partciph_s_dict or cid not in sign_cps_dict:
                ver_sign_cps_res = (False,cid)
                break
            di_inx = self.dis[cid].di_index
            partciph_s = partciph_s_dict[cid]
            sign_cps = sign_cps_dict[cid]
            ver_sign_cps_dict[cid], msg_ciph_s_dict[cid] = sv.verCiphsSign(di_inx, partciph_s, sign_cps, return_msg_flag=True)
            if not ver_sign_cps_dict[cid]:
                ver_sign_cps_res = (False,cid)
                break
        end_time = time.time()
        time_ver_sign_cps = round(end_time-begin_time, 4)
        # stats.update( { 'time_ver_sign_cps':round(end_time-begin_time, 4)} )
        print('  ver_sign_cps: ', dict(ver_sign_cps_dict))
        
        if ver_sign_cps_res[0]:
            # Verify signatures of local models
            begin_time = time.time()
            for cid in sign_lm_dict.keys():
                di_inx = self.dis[cid].di_index
                sign_hlm, msg_hlm = sign_lm_dict[cid]
                ver_sign_lm_dict[cid] = sv.verSolnModelSign(di_inx, msg_hlm, sign_hlm)
            end_time = time.time()
            time_ver_sign_lm = round(end_time-begin_time, 4)
            # stats.update( { 'time_ver_sign_lm':round(end_time-begin_time, 4)} )
            print('  ver_sign_lm: ', dict(ver_sign_lm_dict))
        
        time_ver_local_signs = time_ver_sign_cps + time_ver_sign_lm
        ver_local_sign_dict = {}
        for cid, ver_sign_lm in ver_sign_lm_dict.items():
            di_inx = self.dis[cid].di_index
            ver_local_sign_dict[cid] = ver_sign_lm and ver_sign_cps_dict[cid]

        stats = {   'time_ver_sign_cps': time_ver_sign_cps, 
                    'time_ver_sign_lm': time_ver_sign_lm, 
                    'time_ver_local_signs': time_ver_local_signs } 
                    # 'bytes_msg_ciph_s_dict':sys.getsizeof(msg_ciph_s_dict) }
        print('  >>> ver_local_sign_dict: {}, using time: {}'.format(dict(ver_local_sign_dict),time_ver_local_signs))

        if eval_mem_flag: # statistics memory size
            msg_ciph_s_dict_data = {}
            for k,lhsmsgs in msg_ciph_s_dict.items():
                msg_ciph_s_dict_data[k] = lhsmsgs.getMsgsStr()
            sv_dir = mkdir(self.local_data_dir+'SV/')
            sv_file_path = sv_dir + 'msg_ciph_s_dict_data.pkl'
            self.trainer_save_data(msg_ciph_s_dict_data, file_path=sv_file_path, print_flag=True)
            stats.update({'bytes_msg_ciph_s_dict': os.path.getsize(sv_file_path)})

        return ver_sign_cps_res, ver_local_sign_dict, msg_ciph_s_dict, stats



    def aggregation_partciph_iterkey(self, partciph_s_dict, sign_cps_dict, eval_mem_flag=False):
        ''' SV: 
                1. ciphAggregate(partciph_s_dict): aggregates the ciphertexts of iterkey
                2. evalSigns(sign_cps_dict): genaretes the signature of iterkey ciphertext
                    - i.e., sign(ciph_s_sum) = MKLHS.eval(sign_cps_dict)]
                
                output: ciph_s_sum, sign_ciph_s_sum, stats
                - ciph_s_sum = ciphAggregate(partciph_s_list)
                - sign_ciph_s_sum = evalSigns(sign_cps_list, weights=None)
                - stats includes 'time_agg_ciph_s', 'time_sign_ciph_s_sum', 
                                'bytes_agg_ciph_s', 'bytes_sign_ciph_s_sum'
        '''
        sv = self.svs[0]
        eval_mem_flag = eval_mem_flag if eval_mem_flag else self.eval_mem_flag
        print('(SV) Aggregate & Sign ciphertexts of iterkey:')
        
        # aggregate ciphertexts
        if self.clients_weights:
            partciph_s_list, sign_cps_list, weights = [],[],[]
            for cid, partciph_s in partciph_s_dict.items():
                partciph_s_list.append(partciph_s)
                sign_cps_list.append(sign_cps_dict[cid])
                weights.append(self.clients_weights[cid])
        else:
            partciph_s_list = list(partciph_s_dict.values())
            sign_cps_list = list(sign_cps_dict.values())
            weights = None

        begin_time = time.time()
        ciph_s_sum = sv.ciphAggregate(partciph_s_list, weights=weights)
        time_agg_ciph_s = round(time.time()-begin_time, 4)

        # genarete the signature of iterkey ciphertext
        begin_time = time.time()
        sign_ciph_s_sum = sv.evalSigns(sign_cps_list, weights=weights)
        end_time = time.time()
        time_sign_ciph_s_sum = round(end_time-begin_time, 4)

        stats = {   'time_agg_ciph_s':time_agg_ciph_s, 
                    'time_sign_ciph_s_sum':time_sign_ciph_s_sum }
                    # 'bytes_agg_ciph_s':sys.getsizeof(ciph_s_sum),
                    # 'bytes_sign_ciph_s_sum':sys.getsizeof(sign_ciph_s_sum)}

        if eval_mem_flag: # statistics memory size
            sign_ciph_s_sum_data = sign_ciph_s_sum.getSignStr()
            sv_dir = mkdir(self.local_data_dir+'SV/')
            sv_file_path = sv_dir + 'sign_ciph_s_sum_data.pkl'
            self.trainer_save_data(sign_ciph_s_sum_data, file_path=sv_file_path, print_flag=True)
            stats.update({'bytes_sign_ciph_s_sum': os.path.getsize(sv_file_path)})
        
        return ciph_s_sum, sign_ciph_s_sum, stats


    def ver_sign_ciph_iterkey(self, sign_ciph_s_sum, ciph_s_sum, msg_ciph_s_dict):
        ''' each online DI: (only simulating DIs[0])
                verifies SV's signature of ciph_iterkey, 
                i.e. verifies (sign_ciph_s_sum, ciph_s_sum, msg_ciph_s_dict)
                    - ver1: msg(ciph_s_sum) == Agg(msg_ciph_s_dict)
                    - ver2: verMKSign(sign_ciph_s_sum, msg_ciph_s_dict)
            output:
                - ver_sign_ciph_s_sum = True if ver1&&ver2 else False
                - stat includes 'time_ver_sign_ciph_iterkey'
        '''
        print('(DIs) Verify signature of sign_ciph_s_sum:')
        print('    (only simulating DIs[0])', end='', flush=True)
        # print('sign_ciph_s_sum: ', sign_ciph_s_sum)
        # print('ciph_s_sum: ', ciph_s_sum)
        # print('msg_ciph_s_dict: ', msg_ciph_s_dict)
        
        di = self.dis[0]
        stats = defaultdict(dict) # stats
        
        di_inxs, msgs_ciph_s, weights = [], [], []
        for cid, msg_ciph_s in msg_ciph_s_dict.items():
            di_inxs.append( self.dis[cid].di_index )
            msgs_ciph_s.append( msg_ciph_s )
        pks = [ di.get_di_sign_public_key(di_inx) for di_inx in di_inxs ]
        
        if self.clients_weights:
            weights = [ self.clients_weights[cid] for cid in msg_ciph_s_dict.keys() ]
        else:
            weights = None
        # print("ver_sign_ciph_iterkey weights:", weights)

        begin_time = time.time()
        ver_sign_ciph_s_sum = di.verCiphSumMKSign(ciph_s_sum, sign_ciph_s_sum, pks, msgs_ciph_s, weights=weights)
        time_ver_sign_ciph_iterkey = round(time.time()-begin_time, 4)
        stats[di.di_index].update({'time_ver_sign_ciph_iterkey':time_ver_sign_ciph_iterkey})
        print('\t done! using {} seconds'.format(time_ver_sign_ciph_iterkey))
        print('\t >>> ver_sign_ciph_s_sum:', ver_sign_ciph_s_sum)
        
        return ver_sign_ciph_s_sum, stats


    def local_decshare(self, ciph_s_sum, recovered_clients=None, eval_mem_flag=False):
        ''' each selected DI:
                calculates its partdecshare of ciph_s_sum 
                    - execute only if all selected clients are online
                    - ciph_s_sum = Agg(ciph_s), i.e., Enc(iterkey)
                    
        '''
        print('(online selected DIs) Decshare the aggregation ciphertexts by sk:')
        stats = defaultdict(dict) # stats
        eval_mem_flag = eval_mem_flag if eval_mem_flag else self.eval_mem_flag
        decshare_s_dict = {} # part decshares of enc_serect_key
        ciph_sum_ax = getCiphsAXs(ciph_s_sum)
        recovered_clients = recovered_clients if recovered_clients else self.clients
        for c in tqdm(recovered_clients):
            cid = c.cid
            di = self.dis[cid]
            tqdm.write('\n\033[F >> DI['+str(di.di_index)+'] (cid:'+str(cid)+')'+'\t decrypting ciph_iterkey...' )
            begin_time = time.time()
            partdecshare_s = di.partDecrypt(ciph_sum_ax)
            end_time = time.time()
            decshare_s_dict[cid] = partdecshare_s
            stats[cid].update({ 'time_decshare_s':round(end_time-begin_time, 4)})
                                # 'bytes_ciph_sum_ax':sys.getsizeof(ciph_sum_ax),
                                # 'bytes_decshare_s':sys.getsizeof(partdecshare_s)})
            tqdm.write('\033[F', end='\r') # remove the prompt in tqdm

        if eval_mem_flag: # statistics memory size
            cid = recovered_clients[0].cid # sample client
            di = self.dis[cid] 
            ciph_sum_ax_data = DataZZ(ciph_sum_ax)
            di_dir = mkdir(self.local_data_dir+'DI/')
            di_file_path = di_dir + 'ciph_sum_ax_data.pkl'
            self.trainer_save_data(ciph_sum_ax_data, file_path=di_file_path, print_flag=True)
            stats[cid].update({'bytes_ciph_sum_ax': os.path.getsize(di_file_path)})

            decshare_s_data = DataZZ(decshare_s_dict[cid])
            di_dir = mkdir(self.local_data_dir+'DI/')
            di_file_path = di_dir + 'decshare_s_data.pkl'
            self.trainer_save_data(decshare_s_data, file_path=di_file_path, print_flag=True)
            stats[cid].update({'bytes_decshare_s': os.path.getsize(di_file_path)})

        return decshare_s_dict, stats


    def local_decshare_sss(self, ciph_s_sum, online_clients, recovered_clients=None, eval_mem_flag=False):
        ''' each online DI: 
                calculates its partdecshare of ciph_s_sum
                    - execute if some selected clients are drop-out
                    - ciph_s_sum = Agg(ciph_s), i.e., Enc(iterkey)
        '''
        print('(online DIs) Decshare the aggregation ciphertexts by di_sk_shares:')
        stats = defaultdict(dict) # stats
        eval_mem_flag = eval_mem_flag if eval_mem_flag else self.eval_mem_flag
        decshare_sss_dict = {} # part decshares of enc_serect_key
        recovered_clients = recovered_clients if recovered_clients else self.clients
        recovered_di_inxs = [ self.dis[c.cid].di_index for c in recovered_clients ]
        ciph_sum_ax = getCiphsAXs(ciph_s_sum)
        for c in tqdm(online_clients[:self.di_threshold]):
            cid = c.cid
            di = self.dis[cid]
            begin_time = time.time()
            partdecshare_sss = di.partAggDecshare(ciph_sum_ax, di_inxs=None)
            end_time = time.time()
            decshare_sss_dict[cid] = partdecshare_sss
            stats[cid].update({ 'time_decshare_s_sss':round(end_time-begin_time, 4)
                                # 'bytes_ciph_sum_ax':sys.getsizeof(ciph_sum_ax),
                                # 'bytes_decshare_s_sss':sys.getsizeof(partdecshare_sss)
                                })

        if eval_mem_flag: # statistics memory size
            cid = recovered_clients[0].cid # sample client
            di = self.dis[cid] 
            ciph_sum_ax_data = DataZZ(ciph_sum_ax)
            di_dir = mkdir(self.local_data_dir+'DI/')
            di_file_path = di_dir + 'ciph_sum_ax_data.pkl'
            self.trainer_save_data(ciph_sum_ax_data, file_path=di_file_path, print_flag=True)
            stats[cid].update({'bytes_ciph_sum_ax': os.path.getsize(di_file_path)})

            decshare_sss_data = DataUint64_t(decshare_sss_dict[cid])
            di_dir = mkdir(self.local_data_dir+'DI/')
            di_file_path = di_dir + 'decshare_sss_data.pkl'
            self.trainer_save_data(decshare_sss_data, file_path=di_file_path, print_flag=True)
            stats[cid].update({'bytes_decshare_s_sss': os.path.getsize(di_file_path)})

        return decshare_sss_dict, stats


    def get_iterkey(self, ciph_s_sum, decshare_s_dict, sss_flag=False, eval_mem_flag=False):
        ''' SV: 
                get the aggregation decrypt key 
        '''
        print('(SV) Get iterkey: ')
        eval_mem_flag = eval_mem_flag if eval_mem_flag else self.eval_mem_flag
        # dec_mes = []
        begin_time = time.time()
        if not sss_flag:
            decshare_s_list = list(decshare_s_dict.values())
        else:
            print(' - Recovering decshare for iterkey: ')
            decshare_s_list = self.svs[0].partAggDecshareRecover(decshare_s_dict)

        iterkey = self.svs[0].getIterkey(ciph_s_sum, decshare_s_list)
        time_get_iterkey = round(time.time()-begin_time, 4)
        stats = {   'time_get_iterkey':time_get_iterkey }
        print('\t done! using {} seconds'.format(time_get_iterkey))

        if eval_mem_flag: # statistics memory size
            iterkey_data = DataZZ(iterkey)
            sv_dir = mkdir(self.local_data_dir+'SV/')
            sv_file_path = sv_dir + 'iterkey_data.pkl'
            self.trainer_save_data(iterkey_data, file_path=sv_file_path, print_flag=True)
            stats.update({'bytes_iterkey': os.path.getsize(sv_file_path)})

        return iterkey, stats


    def get_globalmodel(self, ciph_lm_dict, sign_lm_dict, iterkey=None, eval_mem_flag=False):   
        ''' SV: 
                1. ciph_gm = ciphModelAggregate(ciph_lm_dict)
                2. gm = decByIterkey(ciph_gm, iterkey)
                3. sign_gm = MKLHS.eval(sign_lm_dict)
                    - i.e., sign(LHH(gm)) = MKLHS.eval(sign_lm_dict)

            output: gm, sign_gm, hash_lm_dict, stats
                - gm: aggregation of local models
                - sign_gm: signature of gm
                - hash_lm_dict = {cid: msg_ciph_s}
                - stats includes 'time_get_ciph_gm', 'time_get_gm_soln', 'time_sign_gm',
                        'bytes_gm_soln', 'bytes_sign_gm', 'bytes_hash_lm_dict'
        '''
        sv = self.svs[0]
        eval_mem_flag = eval_mem_flag if eval_mem_flag else self.eval_mem_flag

        # get weights of remain local models
        if not self.clients_weights:
            weight_dict, weights = None, None
        else:
            weight_dict = {}
            for cid in ciph_lm_dict.keys():
                di_inx = self.dis[cid].di_index
                weight_dict[di_inx] = self.clients_weights[cid]
            weights = list(weight_dict.values())
        
        # change keys of ciph_lm_dict from cid to di_inx
        ciph_lm_di_dict = {}
        for cid, ciph in ciph_lm_dict.items():
            di_inx = self.dis[cid].di_index
            ciph_lm_di_dict[di_inx] = ciph
        ciph_lm_dict = ciph_lm_di_dict

        # aggregate local model ciphertexts
        print('(SV) Get global model:')
        print(' - Aggregating ciphertexts of global model...', end='', flush=True)
        begin_time = time.time()
        ciph_gm = sv.ciphModelAggregate(ciph_lm_dict, weight_dict=weight_dict)
        end_time = time.time()
        time_get_ciph_gm = round(end_time-begin_time, 4)
        print('\t done! using {} seconds'.format(time_get_ciph_gm))
        
        # aggregate local model ciphertexts
        print(' - Decrypting global model by iterkey...', end='', flush=True)
        begin_time = time.time()
        # gm = sv.getGlobalModelByIterkey(ciph_gm, self.model, iterkey=iterkey, model_name=self.model_name)
        gm = sv.getGlobalModelByIterkey(ciph_gm, self.model_len, iterkey=iterkey)
        end_time = time.time()
        time_get_gm_soln = round(end_time-begin_time, 4)
        print('\t done! using {} seconds'.format(time_get_gm_soln))

        # Genarete the signature of global model by MKLHS.eval(sign_lm_dict)
        print(' - Signing global model...', end='', flush=True)
        sign_lm_list, hash_lm_dict = [], {}
        for cid, sign_lm in sign_lm_dict.items():
            sign_lm_list.append(sign_lm[0])
            hash_lm_dict[cid] = sign_lm[1]
        sign_gm = sv.evalSigns(sign_lm_list, weights=weights)
        end_time = time.time()
        time_sign_gm = round(end_time-begin_time, 4)
        print('\t done! using {} seconds'.format(time_sign_gm))


        stats = {   'time_get_ciph_gm': time_get_ciph_gm,
                    'time_get_gm_soln': time_get_gm_soln,
                    'time_sign_gm': time_sign_gm,
                    'bytes_gm_soln': sys.getsizeof(gm),
                    # 'bytes_sign_gm':sys.getsizeof(sign_gm),
                    # 'bytes_hash_lm_dict':sys.getsizeof(hash_lm_dict) 
                }

        if eval_mem_flag: # statistics memory size
            sign_gm_data = sign_gm.getSignStr()
            sv_dir = mkdir(self.local_data_dir+'SV/')
            sv_file_path = sv_dir + 'sign_gm_data.pkl'
            self.trainer_save_data(sign_gm_data, file_path=sv_file_path, print_flag=True)
            stats.update({'bytes_sign_gm': os.path.getsize(sv_file_path)})

            hash_lm_dict_data = {}
            for k,lhhhash in hash_lm_dict.items():
                hash_lm_dict_data[k] = LHH_init().getHashesHex(lhhhash)
            sv_dir = mkdir(self.local_data_dir+'SV/')
            sv_file_path = sv_dir + 'bytes_hash_lm_dict_data.pkl'
            self.trainer_save_data(hash_lm_dict_data, file_path=sv_file_path, print_flag=True)
            stats.update({'bytes_hash_lm_dict': os.path.getsize(sv_file_path)})

        return gm, sign_gm, hash_lm_dict, stats


    def ver_sign_globalmodel(self, gm, sign_gm, hash_lm_dict):
        ''' each DI: 
                1. ver_sign_gm = verGlobalModelMKSign(gm, sign_gm, hash_lm_dict, pks, hash_lms)
                    - ver1: hash(gm) == agg(hash(lm))
                    - ver2: verMKSign(sign_gm, hash_lm_dict)
                2. gm_soln = (gm/self.soln_scaling)/sum(weights)

            output: gm_soln, sign_gm, hash_lm_dict, stats
                - gm_soln: aggregation of local models
                - sign_gm: signature of gm_soln
                - hash_lm_dict = {cid: hash_lm}
                - stats includes 'time_get_ciph_gm', 'time_get_gm_soln', 'time_sign_gm',
                        'bytes_gm_soln', 'bytes_sign_gm', 'bytes_hash_lm_dict'
        '''
        print('(DIs) Verify signature of sign_gm_soln:')
        print('    (only simulating DIs[0])', end='', flush=True)
        di = self.dis[0]
        stats = defaultdict(dict) # stats

        di_inxs, hash_lms = [], []
        for cid, msg_ciph_s in hash_lm_dict.items():
            di_inxs.append( self.dis[cid].di_index )
            hash_lms.append( msg_ciph_s )
        pks = [ di.get_di_sign_public_key(di_inx) for di_inx in di_inxs ]

        weights = None
        if self.clients_weights:
            weights = [self.clients_weights[cid] for cid in hash_lm_dict.keys()]

        begin_time = time.time()
        ver_sign_gm = di.verGlobalModelMKSign(self.model, gm, sign_gm, pks, hash_lms, weights=weights)
        time_ver_sign_gm = round(time.time()-begin_time, 4)
        stats[di.di_index].update({ 'time_ver_sign_gm': time_ver_sign_gm})
        print('\t done! using {} seconds'.format(time_ver_sign_gm))
        print('\t >>> ver_sign_gm:', ver_sign_gm)

        sum_weights = len(hash_lm_dict) if not weights else sum(weights)
        gm_soln = (torch.tensor(gm)/sum_weights)/self.soln_scaling
        
        return gm_soln, ver_sign_gm, stats


    def train(self):
        ''' The whole training procedure
            No returns. All results all be saved.
        '''
        raise NotImplementedError


    def orgstats2logstats(self, stats):
        logstat = {}
        for stat in stats:
            cid = stat.pop('id')
            logstat[cid] = stat
            stat['id'] = cid
        return logstat


    def log_stats(self, round_i, selected_clients, stats_dict, file_path=None):
        ''' log stats to txt '''
        file_path = mkdir(self.stat_dir)+'/stats.txt' if not file_path else file_path
        stat_str = ''

        if round_i == 0:
            stat_str += '\n================ ' + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + ' ================\n'
            # if self.dataset_name == 'randomvector':
            stat_str += 'Input length: '+str(self.model_len)+'\n'

        stat_str += 'Round ['+str(round_i+1)+']\n'
        if self.clients_weights:
            client_info = [(c.cid, self.clients_weights[c.cid]) for c in selected_clients]
            stat_str += 'clients: '+str(sorted(client_info))+'\n'
        with open(file_path, 'a') as f:
            for stats_name,stats in stats_dict.items():
                stat_str += '>> '+stats_name+' <<\n'
                if type(stats) is type({}) or type(stats) is type(defaultdict()):
                    for k,v in sorted(stats.items()):
                        if type(v) is type({}) or type(v) is type(defaultdict()):
                            stat_str += '  role['+str(k)+'] '
                            for kk,vv in v.items():
                                stat_str += '\n     - '+str(kk)+': '+str(vv)
                        else:
                            stat_str += '  - ' +str(k)+': '+str(v)
                        stat_str += '\n'
                else:
                    stat_str += '  '+str(stats)+'\n'
            f.write(stat_str)
        f.close()

    def log_stats_sheet(self, round_i, stats_dict, file_path=None):
        ''' log stats to sheet '''
        file_path = mkdir(self.stat_dir)+'stats.xlsx' if not file_path else file_path
        try:
            workbook = openpyxl.load_workbook(file_path)
        except:
            workbook = openpyxl.Workbook()

        # transfer stats_dict to DIs' stats_dicts & SV(s)' stats_dicts
        stats_dict_dis = defaultdict(dict)
        stats_dict_svs = defaultdict(dict)
        for stat_name, stats in stats_dict.items():
            if 'DI' in stat_name:
                for k, stat in stats.items():
                    if type(stat) is type({}) or type(stat) is type(defaultdict()):
                        stats_dict_dis[k].update(stat)
                    else:
                        stats_dict_dis[-1][k] = stat
            elif 'SV' in stat_name:
                for k, stat in stats.items():
                    if type(stat) is type({}) or type(stat) is type(defaultdict()):
                        stats_dict_svs[k].update(stat)
                    else:
                        stats_dict_svs[-1][k] = stat
            else:
                stats_dict_svs[-1][stat_name] = stats
        dis_stat_dict = defaultdict(list)
        for stats in stats_dict_dis.values():
            for key,value in sorted(stats.items()):
                dis_stat_dict[key].append(value)

        svs_stat_dict = defaultdict(list)
        for stats in stats_dict_svs.values():
            for key,value in sorted(stats.items()):
                svs_stat_dict[key].append(value)

        variable_dict = {   'round_i': round_i,
                            'drop': self.dropout_flag,
                            'clients_per_round': self.clients_per_round,
                            'di_threshold': self.di_threshold,
                            'model_len': self.model_len
                    }
        variable_keys = sorted(list(variable_dict.keys()))

        # add new row for DI_AVG
        stat_keys = sorted(list(dis_stat_dict.keys()))
        keys = variable_keys + stat_keys
        try:
            sheet = workbook['DI_AVG']
        except:
            sheet = workbook.create_sheet('DI_AVG')
            sheet.append(keys)
        row_datas = [ variable_dict[k] for k in variable_keys ]
        for key in stat_keys:
            values = numpy.array(dis_stat_dict[key])
            avg_value = sum(values)/len(values)
            row_datas.append(avg_value)
        sheet.append(row_datas)

        # add new row for SV_AVG
        stat_keys = sorted(list(svs_stat_dict.keys()))
        keys = variable_keys + stat_keys
        try:
            sheet = workbook['SV_AVG']
        except:
            sheet = workbook.create_sheet('SV_AVG')
            sheet.append(keys)
        row_datas = [ variable_dict[k] for k in variable_keys ]
        for key in stat_keys:
            values = numpy.array(svs_stat_dict[key])
            if len(values) == 1:
                row_datas.append(values[0])
            else:
                avg_value = sum(values)/len(values)
                row_datas.append(avg_value)
        sheet.append(row_datas)

        workbook.save(file_path)
        print('  >>> save stats to sheet ['+str(file_path)+']')
    

    def trainer_save_data(self, data, data_name=None, file_path=None, print_flag=False):
        if not file_path:
            dir_path = mkdir(self.setup_data_dir)
            file_path = str(dir_path)+str(data_name)+'_data.pkl'
        saveData(data, data_name=data_name, file_path=file_path, print_flag=print_flag)


    def trainer_get_data(self, data_name=None, file_path=None, print_flag=False):
        if not file_path:
            dir_path = self.setup_data_dir
            file_path = str(dir_path)+str(data_name)+'_data.pkl'
        return getData(data_name=data_name, file_path=file_path, print_flag=print_flag)


    def get_partciph_s_dict_from_file(self, selected_clients, round_num='tmp'):
        dir_path = self.local_data_dir
        partciph_s_dict = {}
        for c in selected_clients:
            di_inx = self.dis[c.cid].di_index
            di_file_path = dir_path+'DI_'+str(di_inx)+'/round_' + str(round_num) + '_ciph_s_data.pkl'
            data = self.trainer_get_data(file_path=di_file_path, print_flag=True)
            partciph_s = data2Obj(data, 'Ciphertexts', print_flag=True)
            partciph_s_dict[c.cid] = partciph_s
        return partciph_s_dict

    def get_ciph_lm_dict_from_file(self, selected_clients, round_num='tmp'):
        dir_path = self.local_data_dir
        ciph_lm_dict = {}
        for c in selected_clients:
            di_inx = self.dis[c.cid].di_index
            di_file_path = dir_path+'DI_'+str(di_inx)+'/round_' + str(round_num) + '_ciph_lm_data.pkl'
            data = self.trainer_get_data(file_path=di_file_path, print_flag=True)
            ciph_lm = data2Obj(data, 'CiphModel', print_flag=True)
            ciph_lm_dict[c.cid] = ciph_lm
            # ciph_lm_dict[c.cid] = self.trainer_get_data(file_path=di_file_path, print_flag=True)
        return ciph_lm_dict


    def select_clients(self, seed=1):
        """ Selects num_clients clients weighted by number of samples from possible_clients
        Args:
            1. seed: random seed
            2. num_clients: number of clients to select; default 20
                note that within function, num_clients is set to min(num_clients, len(possible_clients))
        Return:
            list of selected clients objects
        """
        num_clients = min(self.clients_per_round, len(self.clients))
        np.random.seed(seed)
        selected_clients = np.random.choice(self.clients, num_clients, replace=False).tolist()
        return selected_clients

    def simulation_dropout(self, seed=1, selected_ids=None, threshold=None, dropout_ratio=None):
        ''' '''
        print('Dropout Simulation:')
        selected_ids = selected_ids if selected_ids else [ c.cid for c in self.clients ]
        threshold = threshold if threshold else self.di_threshold
        dp_ratio = dropout_ratio if dropout_ratio else self.dropout_ratio
        # print('dropout_ratio:', dp_ratio)
        random.seed(seed)

        print(' - all ids:', selected_ids)
        max_dp_num = max( 0, self.di_num-threshold )
        dp_num = min( max_dp_num, ceil(dp_ratio*len(selected_ids)) )
        #print(' - dropout number:', dp_num)
        dp_ids = random.sample( selected_ids, dp_num )
        print(' - dropout ids:', dp_ids)
        return dp_ids


    def simulation_evafailed(self, selected_clients, seed=1, threshold=None, dp_clients_ids=[], failed_ratio=1):
        ''' '''
        print('Evaluation-failed Simulation:')
        selected_cids = [ c.cid for c in selected_clients if c.cid not in dp_clients_ids ]
        threshold = threshold if threshold else self.di_threshold
        random.seed(seed)
        print(' - dropout clients:', dp_clients_ids)
        print(' - online selected_clients:', selected_cids)

        max_fail_num = max( 0, len(selected_cids)-threshold )
        fail_num = min( max_fail_num, floor(failed_ratio*len(selected_cids)) )
        #print(' - eval-failed number:', fail_num)
        fail_clients_ids = random.sample( selected_cids, fail_num )
        print(' - eval-failed clients:', fail_clients_ids)
        return fail_clients_ids


